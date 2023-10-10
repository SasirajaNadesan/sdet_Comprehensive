import openpyxl


#loading the excelsheet
wb = openpyxl.load_workbook("ReaDWriteData.xlsx")
#passing the sheet
sh1=wb['Sheet1']

#counting the Max row and columns
row=sh1.max_row
column = sh1.max_column

#for loop to read the data
for i in range(1,row+1):
    for j in range(1,column+1):
        print(sh1.cell(i,j).value,end='\t')
        print('\n')